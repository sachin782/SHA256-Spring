"""Excel report writer for multi-file output."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from models import ChecksumRecord

WORKSHEET_NAME = "Checksums"
HEADERS = ("File Name", "File Path", "SHA256")
_INVALID_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")
_MAX_SHEET_NAME_LEN = 31


def write_excel_report(records: Sequence[ChecksumRecord], output_path: Path) -> None:
    """Write checksum results to a single-sheet Excel workbook."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = WORKSHEET_NAME
    _populate_sheet(sheet, records)
    workbook.save(output_path)


def write_excel_report_tabbed(
    archive_groups: Mapping[str, Sequence[ChecksumRecord]],
    other_records: Sequence[ChecksumRecord],
    output_path: Path,
) -> None:
    """Write checksum results with one worksheet per archive.

    Args:
        archive_groups: Mapping of archive filename to checksum rows for that archive.
        other_records: Non-archive files placed on the default Checksums sheet.
        output_path: Destination workbook path.
    """
    workbook = Workbook()
    used_names: set[str] = set()

    if other_records:
        sheet = workbook.active
        sheet.title = _unique_sheet_name(WORKSHEET_NAME, used_names)
        _populate_sheet(sheet, other_records)
    else:
        workbook.remove(workbook.active)

    for archive_name, records in archive_groups.items():
        sheet = workbook.create_sheet(title=_unique_sheet_name(archive_name, used_names))
        _populate_sheet(sheet, records)

    if not workbook.sheetnames:
        sheet = workbook.create_sheet(title=WORKSHEET_NAME)
        _populate_sheet(sheet, [])

    workbook.save(output_path)


def _populate_sheet(sheet: Worksheet, records: Sequence[ChecksumRecord]) -> None:
    """Fill a worksheet with formatted checksum rows."""
    for col_index, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(row=1, column=col_index, value=header)
        cell.font = Font(bold=True)

    for row_index, record in enumerate(records, start=2):
        sheet.cell(row=row_index, column=1, value=record.file_name)
        sheet.cell(row=row_index, column=2, value=record.file_path)
        value = record.sha256 if record.success else f"ERROR: {record.error_message}"
        sheet.cell(row=row_index, column=3, value=value)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:C{max(1, len(records) + 1)}"

    for col_index in range(1, 4):
        column = get_column_letter(col_index)
        max_length = 0
        for cell in sheet[column]:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[column].width = min(max(max_length + 2, 12), 80)


def _sanitize_sheet_name(name: str) -> str:
    """Return an Excel-safe worksheet title derived from an archive filename."""
    cleaned = _INVALID_SHEET_CHARS.sub("_", name.strip())
    cleaned = cleaned.strip("'") or WORKSHEET_NAME
    return cleaned[:_MAX_SHEET_NAME_LEN]


def _unique_sheet_name(name: str, used_names: set[str]) -> str:
    """Return a unique, Excel-safe worksheet name."""
    base = _sanitize_sheet_name(name)
    candidate = base
    counter = 1
    while candidate in used_names:
        suffix = f"_{counter}"
        candidate = f"{base[: _MAX_SHEET_NAME_LEN - len(suffix)]}{suffix}"
        counter += 1
    used_names.add(candidate)
    return candidate
